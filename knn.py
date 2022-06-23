from openpyxl import Workbook, load_workbook
from pathlib import Path

# Paths for file processing
FOLDER_PATH = Path(__file__).parent
EXCEL_PATH = FOLDER_PATH/"traintest.xlsx"
OUTPUT_PATH = FOLDER_PATH/"kNN_result.xlsx"

# Worksheets name
TRAINING_WS = "train"
TESTING_WS = "test"

def main():
    # Loading excel file
    wb = load_workbook(EXCEL_PATH)
    train = wb[TRAINING_WS]
    test = wb[TESTING_WS]

    # Getting data
    train_rows = train.rows
    test_rows = test.rows
    header = (cell.value for cell in next(train_rows))
    next(test_rows)

    # Creating training data set from excel
    train_set = [Data(id.value, x1.value, x2.value, x3.value, value = y.value) for id, x1, x2, x3, y in train_rows]
    test_set = [Data(id.value, x1.value, x2.value, x3.value) for id, x1, x2, x3, _ in test_rows]

    # Closing excel file
    wb.close()
    
    # Initializing K value using user input
    k = 0
    while k not in range(1, len(train_set)):
        try:
            k = int(input("k value: "))
        except ValueError:
            k = 0

    # kNN
    for test_data in test_set:
        test_data.value = kNN(test_data, train_set, k)

    # Creating output file
    wb = Workbook()
    ws = wb.active
    ws.title = "Output kNN"
    ws.append(header)
    
    for data in test_set:
        ws.append(data.get_row())

    wb.save(OUTPUT_PATH)


class Data():
    '''
        Data class to represent data in traintest.xlsx
    '''
    def __init__(self, id,  *args, value=None):
        self.id = id
        self.value = value
        self.coords = []
        for point in args:
            self.coords.append(point)

    def get_row(self):
        'returns a tuple containing data in the instance'
        return self.id, *self.coords, self.value

def euclidian_distance(data1, data2):
    '''
        Returns the euclidian distance
    '''
    # Initializing sum value
    sum = 0

    # Adding the squared distance between 2 coordinate in each data to sum
    for point1, point2 in zip(data1.coords, data2.coords):
        sum += (point2-point1)**2

    # Returning square root of sum
    return sum**(1/2)

def kNN(test_data, train_set, k=1):
    '''
        Returns a classification value of test_data based on train_set using kNN algorithm
    '''
    # Listing neighbor. Sorted ascending by euclidian distance
    neighbor_list = sorted(train_set, key=lambda train_data: euclidian_distance(train_data, test_data))

    # Initializing classification value counter
    value_counter = {0:0, 1:0}

    # Counting classification value of first neighbor to k-th neighbor
    for data in neighbor_list[:k]:
        value_counter[data.value]+=1

    # Returns the classification value with the most count 
    return max(value_counter, key=value_counter.get)

if __name__ == "__main__":
    main()